import os
from tqdm import tqdm
from bs4 import BeautifulSoup
import requests
import openpyxl
import mysql.connector as mc

database_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'moody'
}


def create_excel_file(product_name):
    excel = openpyxl.Workbook()
    # print(excel.sheetnames)
    sheet = excel.active
    sheet.title = product_name
    # print(excel.sheetnames)
    sheet.append(['Id', 'Review_Title', 'Review_Comment', 'Rating'])

    return excel, sheet


def add_excel_file_data(sheet, data):
    sheet.append(data)


def save_excel_file(excel, name):
    excel.save('data/' + name)


def get_number_of_pages(soup):
    # Review count
    find_count = soup.find('div', id='reviewContentSection').find('div', id='reviewsSection')
    count = find_count.find('div', class_='review-section-header').h2.text
    count = [int(s) for s in count.split() if s.isdigit()]

    # print(int(count[0]))

    pages = int(int(count[0]) / 10) + 1 if int(count[0]) % 10 != 0 else int(int(count[0]) / 10)
    # print(pages)
    return pages


def get_all_links(link, pages):
    source_link = [link]
    for i in range(pages - 1):
        links = link + '?pgn=' + str(i + 2)
        source_link.append(links)

    # print(source_link)
    return source_link


def url_format(url):
    # get the source url
    source = requests.get(url)
    # validation about the url
    source.raise_for_status()
    # website url format using parser
    soup = BeautifulSoup(source.text, 'html.parser')

    return soup


def get_review_data(section):
    try:
        review_title = section.find('h3', class_='review-item-title').text
        review_comment = section.find('p', class_='review-item-content').text
    # put the value error here if error occur remove value error
    except ValueError:
        review_title = ''
        review_comment = ''
    except AttributeError:
        review_title = ''
        review_comment = ''

    return review_title, review_comment


def get_data(sheet, url, product_id):
    # print(soup)
    data_set = []
    number_of_pages = get_number_of_pages(url_format(url))
    all_links = get_all_links(url, number_of_pages)
    id_num = 0
    for s in all_links:
        soup = url_format(s)
        title = soup.find('div', class_="reviews").find_all('div', class_='ebay-review-section')

        for t in title:
            review_section_left = t.find('div', class_='ebay-review-section-l')
            review_section_right = t.find('div', class_='ebay-review-section-r')
            review_title, review_comment = get_review_data(review_section_right)

            rating = int(review_section_left.find('div', class_='ebay-star-rating').meta['content'])

            # print(review_title,review_comment,rating)
            id_num = id_num + 1
            add_excel_file_data(sheet, [id_num, review_title, review_comment, rating])
            data_set.append([product_id, review_title, review_comment, rating])

    insert_data_into_database(data_set)


def get_data_from_database():
    product_url = []
    product_name = []
    product_id = []
    try:
        conn = mc.connect(**database_config)
        cursor = conn.cursor()
        cursor.execute("USE moody")
        cursor.execute("SELECT * FROM products")
        result = cursor.fetchall()
        for row in result:
            product_url.append(row[3])
            product_name.append(row[2])
            product_id.append(row[1])
        conn.close()
    except mc.Error as e:
        print(e)
    # print(product_url)
    # print(product_name)
    return product_name, product_url, product_id


def insert_data_into_database(data):
    try:
        conn = mc.connect(**database_config)
        cursor = conn.cursor()
        cursor.execute("USE moody")
        insert_query = "INSERT INTO reviews (product_id,review_title,review_comment,rating) VALUES (%s,%s,%s,%s)"
        for i in data:
            cursor.execute(insert_query, i)
        conn.commit()
        conn.close()
    except mc.Error as e:
        print(e)


def database_recreate():
    try:
        conn = mc.connect(**database_config)
        cursor = conn.cursor()
        cursor.execute("USE moody")
        table_name = "reviews"
        drop_query = "DROP TABLE IF EXISTS " + table_name
        cursor.execute(drop_query)
        create_query = ("CREATE TABLE IF NOT EXISTS reviews ("
                        "id INT AUTO_INCREMENT PRIMARY KEY,product_id VARCHAR(255),"
                        "review_title VARCHAR(255),"
                        "review_comment VARCHAR(255),"
                        "rating INT)"
                        )
        cursor.execute(create_query)
        conn.commit()
        conn.close()
    except mc.Error as e:
        print(e)


def main():
    product_name, product_url, product_id = get_data_from_database()
    database_recreate()
    for i in tqdm(range(len(product_name)), desc="Processing products"):
        url = product_url[i]
        file_name = product_name[i]
        excel, sheet = create_excel_file(product_name[i])

        try:
            get_data(sheet, url, product_id[i])
        except Exception as e:
            print(e)
        # check if the file exist or not if file exist
        if file_name + '.xlsx' in os.listdir():
            save_excel_file(excel, file_name + '1.xlsx')
        save_excel_file(excel, file_name + '.xlsx')


if __name__ == '__main__':
    main()
